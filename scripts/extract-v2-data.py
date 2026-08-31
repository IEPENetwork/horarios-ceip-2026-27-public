#!/usr/bin/env python3
"""Build the site data files from the verified V2 workbook and substitution JSON."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "source-v2"
OUTPUT = ROOT / "src" / "data"
WORKBOOK = SOURCE / "horarios_ceip_2026_2027_V2_CON_SUSTITUCIONES.xlsx"
SUBSTITUTIONS = SOURCE / "sustituciones_v2_site_data.json"

DAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
GROUPS = ["1.º", "2.ºA", "2.ºB", "3.º", "4.º", "5.ºA", "5.ºB", "6.ºA", "6.ºB"]
SLOTS = {
    day: ["09:00–10:00", "10:00–10:45", "10:45–11:30", "11:30–12:00", "12:00–13:00", "13:00–14:00"]
    for day in DAYS[:4]
}
SLOTS["Viernes"] = ["09:00–10:30", "10:30–11:15", "11:15–11:45", "11:45–12:30", "12:30–14:00"]
RECESS = {day: "11:30–12:00" for day in DAYS[:4]} | {"Viernes": "11:15–11:45"}


def records(ws, header_row=1):
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row]]
    result = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None for value in row):
            continue
        result.append({headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]})
    return result


def slot(value):
    return str(value).replace("-", "–").replace("—", "–").replace(" ", "")


def split_teachers(value):
    if not value:
        return []
    return [name.strip() for name in re.split(r"\s*[,;+]\s*", str(value)) if name.strip()]


def status_minutes(value):
    match = re.search(r"(\d+)\s*min", str(value or ""))
    return int(match.group(1)) if match else 0


def main():
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    lesson_rows = records(wb["Datos"])
    complementary_rows = records(wb["Complementarias"], 3)
    coordination_rows = records(wb["Coordinación tutor-DC"], 3)
    exception_rows = records(wb["Excepciones"], 3)
    matrix_rows = records(wb["Matriz docentes"])
    substitution_source = json.loads(SUBSTITUTIONS.read_text(encoding="utf-8"))

    tutors = {}
    for row in coordination_rows:
        group = row.get("Grupo")
        tutor = row.get("Tutor/a")
        if group in GROUPS and tutor:
            tutors[group] = tutor

    lessons = []
    source_anomalies = []
    direct_seen = set()
    shared_seen = set()
    direct_minutes = defaultdict(int)
    shared_minutes = defaultdict(int)
    subject_teachers = defaultdict(set)

    for row in lesson_rows:
        group = row["Grupo"]
        day = row["Día"]
        time = slot(row["Franja"])
        minutes = int(row["Minutos"])
        primary = str(row["Docente principal"]).strip()
        raw_shared = split_teachers(row.get("Docencia compartida / apoyo"))
        shared = [teacher for teacher in raw_shared if teacher != primary]
        if len(shared) != len(raw_shared):
            source_anomalies.append({
                "type": "redundant-primary-as-shared",
                "group": group,
                "day": day,
                "time": time,
                "subject": str(row["Asignatura"]).strip(),
                "teacher": primary,
                "resolution": "Se representa una sola vez como docente principal; no se altera la sesión.",
            })
        subject = str(row["Asignatura"]).strip()
        lessons.append({
            "group": group,
            "day": day,
            "time": time,
            "minutes": minutes,
            "subject": subject,
            "primary": primary,
            "shared": shared,
            "notes": row.get("Observaciones") or "",
        })
        primary_key = (primary, day, time)
        if primary_key not in direct_seen:
            direct_seen.add(primary_key)
            direct_minutes[primary] += minutes
        subject_teachers[subject].add(primary)
        for teacher in shared:
            shared_key = (teacher, day, time)
            if shared_key not in shared_seen:
                shared_seen.add(shared_key)
                shared_minutes[teacher] += minutes
            subject_teachers[subject].add(teacher)

    teachers = list(substitution_source["teachers"])
    complements = defaultdict(lambda: defaultdict(int))
    complementary_events = []
    for row in complementary_rows:
        teacher = row.get("Docente")
        concept = row.get("Concepto")
        if teacher not in teachers or not concept:
            continue
        minutes = int(row.get("Total asignado (min)") or 0)
        complements[teacher][concept] += minutes
        complementary_events.append({
            "teacher": teacher,
            "concept": concept,
            "schedule": row.get("Distribución / horario") or "",
            "minutes": minutes,
            "notes": row.get("Observaciones") or "",
        })

    matrix = []
    support_minutes = defaultdict(int)
    for row in matrix_rows:
        day = row["Día"]
        time = str(row["Franja"])
        states = {teacher: str(row.get(teacher) or "") for teacher in teachers}
        matrix.append({"day": day, "time": time, "teachers": states})
        for teacher, state in states.items():
            if state.startswith("Apoyo"):
                support_minutes[teacher] += status_minutes(state)

    teacher_loads = {}
    for teacher in teachers:
        family = complements[teacher]["Atención a familias"]
        coordination = complements[teacher]["Coordinación docente"]
        tutorial = complements[teacher]["Reducción por tutoría"]
        computed = direct_minutes[teacher] + shared_minutes[teacher] + 150 + family + coordination + tutorial
        support = support_minutes[teacher]
        teacher_loads[teacher] = {
            "direct": direct_minutes[teacher],
            "shared": shared_minutes[teacher],
            "recess": 150,
            "family": family,
            "coordination": coordination,
            "tutorial": tutorial,
            "computed": computed,
            "support": support,
            "total": computed + support,
        }

    tutor_names = set(tutors.values())
    shared_names = {teacher for teacher, minutes in shared_minutes.items() if minutes}
    specialist_names = {teacher for teacher in teachers if teacher not in tutor_names and direct_minutes[teacher]}
    support_names = {teacher for teacher, minutes in support_minutes.items() if minutes}
    teacher_roles = {
        teacher: [
            role
            for role, member in (
                ("Tutor/a", teacher in tutor_names),
                ("Especialista", teacher in specialist_names),
                ("Docencia compartida", teacher in shared_names),
                ("Apoyo disponible", teacher in support_names),
            )
            if member
        ]
        for teacher in teachers
    }

    coordinations = []
    for row in coordination_rows:
        if row.get("Grupo") not in GROUPS:
            continue
        coordinations.append({
            "group": row.get("Grupo"),
            "tutor": row.get("Tutor/a"),
            "sharedTeacher": row.get("Docente compartido/apoyo"),
            "time": row.get("Franja semanal coincidente"),
            "tutorType": row.get("Tipo tutor/a"),
            "sharedType": row.get("Tipo docente compartido"),
            "status": row.get("Estado"),
            "notes": row.get("Observaciones"),
        })

    exceptions = [
        {key: value for key, value in row.items() if value is not None}
        for row in exception_rows
        if row.get("ID")
    ]

    output = {
        "version": "V2 final verificada",
        "sourceWorkbook": WORKBOOK.name,
        "days": DAYS,
        "groups": GROUPS,
        "slots": SLOTS,
        "recess": RECESS,
        "tutors": tutors,
        "teachers": teachers,
        "teacherRoles": teacher_roles,
        "subjects": {subject: sorted(names) for subject, names in sorted(subject_teachers.items())},
        "lessons": lessons,
        "teacherMatrix": matrix,
        "teacherLoads": teacher_loads,
        "complementaryEvents": complementary_events,
        "coordinations": coordinations,
        "exceptions": exceptions,
        "sourceAnomalies": source_anomalies,
    }

    OUTPUT.mkdir(parents=True, exist_ok=True)
    (OUTPUT / "schedule-v2.json").write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT / "substitutions-v2.json").write_text(
        json.dumps(substitution_source, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Generated {len(lessons)} lessons, {len(teachers)} teachers and {len(substitution_source['scenarios'])} scenarios")


if __name__ == "__main__":
    main()
